"""Load CSV/XLSX into simple tables and classify their shape.

No pandas: the engine's dependency budget is small and these operations
(parse, transpose, share-of-total) don't need it.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

_NUM_RE = re.compile(r"^\(?-?[\d,]*\.?\d+\)?$")
_PERIOD_RE = re.compile(
    r"(20\d\d|19\d\d|q[1-4]|fy|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
    r"|month|week|quarter|year|ytd|h[12])", re.IGNORECASE)


@dataclass
class Table:
    name: str                       # sheet name or file stem
    source: str                     # file name for provenance
    headers: list[str]
    rows: list[list[str]] = field(default_factory=list)

    def column(self, idx: int) -> list[str]:
        return [row[idx] if idx < len(row) else "" for row in self.rows]


def parse_number(raw) -> float | None:
    """'$1,234.5', '12%', '(340)' -> float; text -> None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    s = str(raw).strip().replace("$", "").replace("%", "").replace("–", "-")
    if not s or not _NUM_RE.match(s):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "")
    try:
        value = float(s)
    except ValueError:
        return None
    return -value if neg else value


def numeric_column_indexes(table: Table) -> list[int]:
    """Columns where at least 70% of non-empty cells parse as numbers."""
    out = []
    for i in range(len(table.headers)):
        cells = [c for c in table.column(i) if str(c).strip() != ""]
        if not cells:
            continue
        parsed = sum(1 for c in cells if parse_number(c) is not None)
        if parsed / len(cells) >= 0.7:
            out.append(i)
    return out


def label_column_index(table: Table, numeric: list[int]) -> int | None:
    for i in range(len(table.headers)):
        if i not in numeric:
            return i
    return None


def looks_like_periods(headers: list[str]) -> bool:
    """True when column headers read as a time axis (Q1 26, Jan, 2024...)."""
    hits = sum(1 for h in headers if _PERIOD_RE.search(str(h)))
    return len(headers) >= 3 and hits / len(headers) >= 0.6


def load_tables(path: Path) -> list[Table]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_xlsx(path)
    raise ValueError(f"insights reads .csv/.xlsx - got '{path.name}'. "
                     "For PDF/DOCX run `deckstudio ingest` and read the digest.")


def _load_csv(path: Path) -> list[Table]:
    text = path.read_text(encoding="utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = [row for row in csv.reader(text.splitlines(), dialect) if any(
        c.strip() for c in row)]
    if len(rows) < 2:
        return []
    return [Table(name=path.stem, source=path.name, headers=rows[0], rows=rows[1:])]


def _load_xlsx(path: Path) -> list[Table]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    tables = []
    for ws in wb.worksheets:
        raw = [["" if v is None else v for v in row]
               for row in ws.iter_rows(values_only=True)]
        raw = [r for r in raw if any(str(c).strip() for c in r)]
        if len(raw) < 2:
            continue
        headers = [str(c).strip() for c in raw[0]]
        rows = [[str(c).strip() if not isinstance(c, (int, float)) else c
                 for c in r] for r in raw[1:]]
        tables.append(Table(name=ws.title, source=path.name,
                            headers=headers, rows=rows))
    wb.close()
    return tables
