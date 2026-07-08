"""XLSX -> markdown digest: every sheet's used range as a table, with
provenance headers so the agent can cite sheet/range for each number."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .digest import Digest, table_to_markdown

MAX_ROWS = 60
MAX_COLS = 20


def digest_xlsx(path: Path) -> Digest:
    digest = Digest(source_name=path.name, kind="xlsx")
    wb = load_workbook(str(path), data_only=True, read_only=True)
    for ws in wb.worksheets:
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            continue
        rows = []
        truncated = False
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= MAX_ROWS:
                truncated = True
                break
            cells = ["" if v is None else _fmt(v) for v in row[:MAX_COLS]]
            if any(c != "" for c in cells):
                rows.append(cells)
        if not rows:
            continue
        digest.add(f"## Sheet: {ws.title} (rows 1-{min(ws.max_row, MAX_ROWS)})\n")
        digest.add(table_to_markdown(rows))
        digest.add("")
        if truncated:
            digest.warn(f"sheet '{ws.title}' truncated at {MAX_ROWS} rows "
                        f"(has {ws.max_row}) — ask for the specific range if more is needed.")
    wb.close()
    return digest


def _fmt(value) -> str:
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)
